from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference


# convertData(value, spreadsheet_type):
# Takes a numerical value and converts it based on the spreadsheet type.
def convertData(value, spreadsheet_type):
    if spreadsheet_type == "temperature":
        return (value - 32) * 5/9
    elif spreadsheet_type == "weight":
        return value / 2.205
    elif spreadsheet_type == "rain":
        return value * 2.54
    else:
        return None


# insertData(path, data):
# Appends a line to a CSV file.
def insertData(path, data):
    try:
        with open(path, "a", encoding="utf-8") as file:
            file.write(data + "\n")
    except Exception as e:
        print("Error writing to file:", e)


# viewData(path):
# Displays CSV file contents.
def viewData(path):
    try:
        with open(path, "r", encoding="utf-8") as file:
            print(file.read())
    except Exception as e:
        print("Error reading file:", e)


# getTempInput():
def getTempInput():
    path = r"C:\PythonFiles\final.csv"
    num = int(input("How many entries are you inputting? "))

    for i in range(num):
        try:
            date = input("Enter date: ")
            value = float(input("Enter temp (F): "))

            converted = convertData(value, "temperature")
            data = f"{date},{value},{converted}"

            insertData(path, data)
            print("Saved:", data)

        except Exception as e:
            print("Error:", e)


# getWeightInput():
def getWeightInput():
    path = r"C:\PythonFiles\final.csv"
    num = int(input("How many entries? "))

    for i in range(num):
        try:
            date = input("Enter date: ")
            value = float(input("Enter weight (lbs): "))

            converted = convertData(value, "weight")
            data = f"{date},{value},{converted}"

            insertData(path, data)
            print("Saved:", data)

        except Exception as e:
            print("Error:", e)


# getRainInput():
def getRainInput():
    path = r"C:\PythonFiles\final.csv"
    num = int(input("How many entries? "))

    for i in range(num):
        try:
            date = input("Enter date: ")
            value = float(input("Enter rain (inches): "))

            converted = convertData(value, "rain")
            data = f"{date},{value},{converted}"

            insertData(path, data)
            print("Saved:", data)

        except Exception as e:
            print("Error:", e)


# createChart(path, chart_type):
# Reads CSV and creates Excel chart.
def createChart(path, chart_type):

    print("Choose data source:")
    print("1 Original Data")
    print("2 Converted Data")
    choice = input("Enter choice: ")

    dates = []
    values = []

    try:
        with open(path, "r", encoding="utf-8") as file:
            for line in file:
                parts = line.strip().split(",")

                if len(parts) == 3:
                    dates.append(parts[0])

                    if choice == "1":
                        values.append(float(parts[1]))
                    else:
                        values.append(float(parts[2]))

    except Exception as e:
        print("Error reading file:", e)
        return

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws["A1"] = "Date"
    ws["B1"] = "Value"

    for i in range(len(dates)):
        ws.cell(row=i+2, column=1, value=dates[i])
        ws.cell(row=i+2, column=2, value=values[i])

    # Create chart
    if chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()

    data = Reference(ws, min_col=2, min_row=1, max_row=len(values)+1)
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(dates)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    chart.x_axis.title = "Date"
    chart.y_axis.title = "Values"
    chart.title = f"carana9896 {datetime.now().strftime('%m/%d/%Y')}"

    ws.add_chart(chart, "D2")

    wb.save(r"C:\PythonFiles\final.xlsx")

    print("Report created successfully!")


# generateReport(path):
# Asks for chart type and calls createChart.
def generateReport(path):

    print("Choose graph type:")
    print("1 Line Chart")
    print("2 Bar Chart")
    choice = input("Enter choice: ")

    if choice == "1":
        createChart(path, "line")
    elif choice == "2":
        createChart(path, "bar")
    else:
        print("Invalid choice.")


# main():
def main():
    path = r"C:\PythonFiles\final.csv"

    print("carana9896's Spreadsheet Automation Menu")
    print("1 Input Data")
    print("2 View Data")
    print("3 Generate Report")

    choice = input("Enter option: ")

    if choice == "1":
        print("1 Temperature\n2 Weight\n3 Rain")
        stype = input("Select type: ")

        if stype == "1":
            getTempInput()
        elif stype == "2":
            getWeightInput()
        elif stype == "3":
            getRainInput()

    elif choice == "2":
        viewData(path)

    elif choice == "3":
        generateReport(path)

    else:
        print("Invalid choice.")


# Run program
main()
